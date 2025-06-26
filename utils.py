"""
utils - 数据分析智能体使用的工具函数

Author: 骆昊
Version: 0.1
Date: 2025/6/25
"""
import json
import pandas as pd
import openpyxl
import io
import streamlit as st
import hashlib

from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from langchain_experimental.agents.agent_toolkits import create_pandas_dataframe_agent
from langchain.callbacks.base import BaseCallbackHandler

class StreamlitCallbackHandler(BaseCallbackHandler):
    """Streamlit流式输出回调处理器"""
    
    def __init__(self, container):
        self.container = container
        self.text = ""
        self.step_count = 0
        
    def on_llm_new_token(self, token: str, **kwargs) -> None:
        """处理新的token"""
        self.text += token
        self.container.markdown(self.text + "▌")
        
    def on_llm_end(self, response, **kwargs) -> None:
        """LLM结束时的处理"""
        self.container.markdown(self.text)
        
    def on_agent_action(self, action, **kwargs) -> None:
        """Agent执行动作时的回调"""
        self.step_count += 1
        self.container.info(f"🔄 执行步骤 {self.step_count}: {action.tool}")
        
    def on_tool_start(self, serialized, input_str: str, **kwargs) -> None:
        """工具开始执行时的回调"""
        tool_name = serialized.get("name", "未知工具")
        self.container.info(f"🛠️ 正在使用工具: {tool_name}")
        
    def on_tool_end(self, output: str, **kwargs) -> None:
        """工具执行完成时的回调"""
        self.container.success(f"✅ 工具执行完成")

PROMPT_TEMPLATE = """你是一位数据分析助手，你的回应内容取决于用户的请求内容，请按照下面的步骤处理用户请求：
1. 思考阶段 (Thought) ：先分析用户请求类型（文字回答/表格/图表），并验证数据类型是否匹配。
2. 行动阶段 (Action) ：根据分析结果选择以下严格对应的格式。
   - 纯文字回答:
     {"answer": "不超过50个字符的明确答案"}

   - 表格数据：
     {"table":{"columns":["列名1", "列名2", ...], "data":[["第一行值1", "值2", ...], ["第二行值1", "值2", ...]]}}

   - 柱状图
     {"bar":{"columns": ["A", "B", "C", ...], "data":[35, 42, 29, ...]}}

   - 折线图
     {"line":{"columns": ["A", "B", "C", ...], "data": [35, 42, 29, ...]}}

   - 散点图
     {"scatter":{"x_data": [1, 2, 3, ...], "y_data": [4, 5, 6, ...], "labels": ["点1", "点2", ...]}}

   - 饼图
     {"pie":{"labels": ["类别1", "类别2", ...], "values": [30, 45, 25, ...]}}

   - 热力图
     {"heatmap":{"data": [[1, 2, 3], [4, 5, 6]], "x_labels": ["A", "B", "C"], "y_labels": ["X", "Y"]}}
     
3. 格式校验要求
   - 字符串值必须使用英文双引号
   - 数值类型不得添加引号
   - 确保数组闭合无遗漏
   错误案例：{'columns':['Product', 'Sales'], data:[[A001, 200]]}
   正确案例：{"columns":["product", "sales"], "data":[["A001", 200]]}

注意：响应数据的"output"中不要有换行符、制表符以及其他格式符号。

当前用户请求如下：\n"""


def generate_cache_key(df, query):
    """生成缓存键"""
    # 使用数据框的形状、列名和查询内容生成唯一键
    # 安全地处理数据类型，避免序列化错误
    try:
        dtype_dict = {col: str(dtype) for col, dtype in df.dtypes.to_dict().items()}
    except Exception:
        dtype_dict = {col: "unknown" for col in df.columns}
    
    df_info = f"{df.shape}_{list(df.columns)}_{dtype_dict}"
    cache_string = f"{df_info}_{query}"
    return hashlib.md5(cache_string.encode()).hexdigest()

@st.cache_data(ttl=3600)  # 缓存1小时
def cached_dataframe_analysis(_df, query, cache_key):
    """缓存的数据分析函数"""
    return _perform_analysis(_df, query)

def _perform_analysis(df, query):
    """执行实际的数据分析"""
    load_dotenv()
    import os
    model = ChatOpenAI(
        base_url='https://oneapi.xty.app/v1',
        api_key=os.getenv("OPENAI_API_KEY"),
        model='gpt-4o-mini',
        temperature=0,
        max_tokens=8192,
        streaming=False  # 关闭流式输出，确保格式一致性
    )
    agent = create_pandas_dataframe_agent(
        llm=model,
        df=df,
        agent_executor_kwargs={"handle_parsing_errors": True},
        max_iterations=32,
        allow_dangerous_code=True,
        verbose=True
    )

    prompt = PROMPT_TEMPLATE + query

    try:
        response = agent.invoke({"input": prompt})
        # 增强JSON解析错误处理
        try:
            return json.loads(response["output"])
        except json.JSONDecodeError:
            # 如果JSON解析失败，返回原始文本作为答案
            return {"answer": response["output"]}
    except Exception as err:
        print(f"分析错误: {err}")
        return {"answer": "暂时无法提供分析结果，请稍后重试！"}

def dataframe_agent(df, query, stream_container=None):
    """数据分析代理函数，支持缓存和流式输出"""
    # 生成缓存键
    cache_key = generate_cache_key(df, query)
    
    # 检查是否启用流式输出
    if stream_container is not None:
        return dataframe_agent_streaming(df, query, stream_container)
    else:
        # 使用缓存
        return cached_dataframe_analysis(df, query, cache_key)

def dataframe_agent_streaming(df, query, stream_container):
    """支持流式输出的数据分析函数"""
    load_dotenv()
    import os
    
    # 创建流式回调处理器
    callback_handler = StreamlitCallbackHandler(stream_container)
    
    model = ChatOpenAI(
        base_url="https://api.openai-hk.com/v1",
        api_key=os.getenv("OPENAI_API_KEY"),
        model="gpt-4o-mini",
        temperature=0,
        max_tokens=8192,
        streaming=False,  # 关闭模型级别的流式输出，避免格式解析问题
        callbacks=[callback_handler]
    )
    agent = create_pandas_dataframe_agent(
        llm=model,
        df=df,
        agent_executor_kwargs={"handle_parsing_errors": True},
        max_iterations=32,
        allow_dangerous_code=True,
        verbose=True
    )

    prompt = PROMPT_TEMPLATE + query

    try:
        # 显示分析开始信息
        stream_container.info("🤖 AI正在分析您的数据...")
        response = agent.invoke({"input": prompt})
        
        # 尝试解析JSON响应
        try:
            result = json.loads(response["output"])
            stream_container.success("✅ 分析完成！")
            return result
        except json.JSONDecodeError as json_err:
            # 如果JSON解析失败，返回原始文本
            stream_container.warning("⚠️ 响应格式解析异常，返回原始结果")
            return {"answer": response["output"]}
            
    except Exception as err:
        error_msg = f"分析过程中出现错误: {str(err)}"
        print(error_msg)
        stream_container.error(error_msg)
        return {"answer": "暂时无法提供分析结果，请稍后重试！"}


def load_data_file(uploaded_file, file_type_option):
    """
    加载不同格式的数据文件
    
    Args:
        uploaded_file: Streamlit上传的文件对象
        file_type_option: 用户选择的文件类型选项
    
    Returns:
        pandas.DataFrame 或包含多个工作表的字典
    """
    file_extension = uploaded_file.name.split('.')[-1].lower()
    
    try:
        if file_type_option.startswith("Excel"):
            # Excel文件处理，添加错误处理
            try:
                wb = openpyxl.load_workbook(uploaded_file)
                if len(wb.sheetnames) > 1:
                    # 多个工作表，返回字典
                    sheets = {}
                    for sheet_name in wb.sheetnames:
                        try:
                            df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                            if not df.empty:  # 只添加非空工作表
                                sheets[sheet_name] = df
                        except Exception as e:
                            print(f"跳过工作表 {sheet_name}: {str(e)}")
                    
                    if not sheets:
                        raise ValueError("所有工作表都无法读取或为空")
                    return {"sheets": sheets}
                else:
                    # 单个工作表
                    return pd.read_excel(uploaded_file)
            except Exception as e:
                raise ValueError(f"Excel文件读取失败: {str(e)}")
                
        elif file_type_option.startswith("CSV"):
            # CSV文件处理，添加编码检测和错误处理
            try:
                return pd.read_csv(uploaded_file, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    return pd.read_csv(uploaded_file, encoding='gbk')
                except UnicodeDecodeError:
                    return pd.read_csv(uploaded_file, encoding='latin-1')
            
        elif file_type_option.startswith("JSON"):
            # JSON文件处理，支持多种JSON结构
            content = uploaded_file.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')
            
            try:
                json_data = json.loads(content)
            except json.JSONDecodeError as e:
                raise ValueError(f"JSON格式错误: {str(e)}")
            
            # 尝试不同的JSON结构
            if isinstance(json_data, list):
                if len(json_data) == 0:
                    raise ValueError("JSON数组为空")
                return pd.DataFrame(json_data)
            elif isinstance(json_data, dict):
                # 检查是否包含数据数组
                for key, value in json_data.items():
                    if isinstance(value, list) and len(value) > 0:
                        if isinstance(value[0], dict):
                            return pd.DataFrame(value)
                
                # 如果是嵌套字典，尝试normalize
                try:
                    return pd.json_normalize(json_data)
                except Exception:
                    # 如果normalize失败，将字典转换为单行DataFrame
                    return pd.DataFrame([json_data])
            else:
                raise ValueError("不支持的JSON格式，请确保JSON包含数组或对象结构")
                
        elif file_type_option.startswith("TSV"):
            # TSV文件处理
            return pd.read_csv(uploaded_file, sep='\t')
            
        elif file_type_option.startswith("Parquet"):
            # Parquet文件处理
            return pd.read_parquet(uploaded_file)
            
        elif file_type_option.startswith("TXT"):
            # TXT文件处理（假设是分隔符分隔的数据）
            content = uploaded_file.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')
            
            # 尝试检测分隔符
            lines = content.strip().split('\n')
            if len(lines) < 2:
                raise ValueError("TXT文件内容不足，无法解析为表格数据")
            
            # 改进的分隔符检测算法
            separators = [',', '\t', ';', '|', ' ']
            best_sep = ','
            max_consistency = 0
            
            # 检查前几行来确定最一致的分隔符
            sample_lines = lines[:min(10, len(lines))]
            
            for sep in separators:
                col_counts = []
                for line in sample_lines:
                    if line.strip():  # 跳过空行
                        col_counts.append(len(line.split(sep)))
                
                if col_counts:
                    # 计算列数的一致性（相同列数的行数）
                    most_common_cols = max(set(col_counts), key=col_counts.count)
                    consistency = col_counts.count(most_common_cols)
                    
                    # 选择一致性最高且列数大于1的分隔符
                    if consistency > max_consistency and most_common_cols > 1:
                        max_consistency = consistency
                        best_sep = sep
            
            # 使用StringIO来模拟文件对象，添加错误处理参数
            string_io = io.StringIO(content)
            try:
                return pd.read_csv(
                    string_io, 
                    sep=best_sep,
                    on_bad_lines='skip',  # 跳过有问题的行
                    engine='python',     # 使用Python引擎，更宽容
                    skipinitialspace=True  # 跳过分隔符后的空格
                )
            except Exception as csv_error:
                # 如果CSV解析失败，尝试作为固定宽度文件处理
                string_io = io.StringIO(content)
                try:
                    return pd.read_fwf(string_io)
                except Exception:
                    raise ValueError(f"无法解析TXT文件格式。原始错误: {str(csv_error)}")
            
        else:
            raise ValueError(f"不支持的文件类型: {file_type_option}")
            
    except Exception as e:
        raise Exception(f"文件解析错误: {str(e)}")
